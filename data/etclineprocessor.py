import openpyxl
import json
import datetime


# 自定义JSON序列化器
class JsonCustomEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime.datetime):
            return obj.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(obj, datetime.date):
            return obj.strftime('%Y-%m-%d')
        else:
            return json.JSONEncoder.default(self, obj)


class DataProcessor(object):
    fields = ('port', 'port_district', 'door_area', 'type', 'start', 'end', 'distance')

    # 无参构造函数
    def __init__(self):
        pass

    # 转成json
    def to_json(self):
        return json.dumps(self, default=lambda obj: obj.__dict__, indent=4, cls=JsonCustomEncoder)


# read xlsx file
# get workbook
sourceWorkbook = openpyxl.load_workbook('../files/鸭嘴兽ETC报价路线导入模板.xlsx')
# get sheet by index
worksheet = sourceWorkbook.worksheets[0]
# get all rows
rows = worksheet.rows
# 遍历所有行
source_data_list = []
heads = []
for row_index, row in enumerate(rows):
    if row_index == 0:
        continue
    if row_index == 1:
        # get values from row
        heads = [cell.value for cell in row]
        continue
    # 遍历所有列, 带索引，从第3行开始
    data = DataProcessor()
    for index, cell in enumerate(row):
        if index == 0:
            data.port = cell.value
        elif index == 1:
            data.port_district = cell.value
        elif index == 2:
            data.door_area = cell.value
        elif index == 3:
            data.type = cell.value
        elif index == 4:
            data.start = str(cell.value)
        elif index == 5:
            data.end = str(cell.value)
        elif index == 6:
            data.distance = str(cell.value)

    source_data_list.append(data)

door_names_sheet = sourceWorkbook.worksheets[1]
door_names = []
# get all rows
door_names_rows = door_names_sheet.rows
# get value from first column every row
for row in door_names_rows:
    door_names.append(row[0].value)

# write source_data_list to new xlsx file
# Create a new workbook
new = openpyxl.Workbook()
# create new sheet
new_sheet = new.create_sheet('Sheet1', 0)
# write data to new sheet
# iterate door_names
new_sheet.append([])
new_sheet.append(heads)
# 前1000个数据
for door_name in door_names:
    for index, data in enumerate(source_data_list):
        new_sheet.append(
            [data.port, data.port_district, door_name, data.type, data.start, data.end,
             data.distance])
# 设置列宽
new_sheet.column_dimensions['A'].width = 20
new_sheet.column_dimensions['B'].width = 20
new_sheet.column_dimensions['C'].width = 20
new_sheet.column_dimensions['D'].width = 20
new_sheet.column_dimensions['E'].width = 20
new_sheet.column_dimensions['F'].width = 20
new_sheet.column_dimensions['G'].width = 20
# set the color of the first row to red
for cell in new_sheet['2:2']:
    cell.font = openpyxl.styles.Font(color='ED7D31')

# save new workbook

new.save('../files/newetcline.xlsx')

sourceWorkbook.close()
