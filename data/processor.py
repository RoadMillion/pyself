import openpyxl
import json
import datetime


# 自定义JSON序列化器
class JsonCustomEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime.datetime):
            return obj.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(obj, datetime.date):
            return obj.strftime('%Y-%m-%d')
        else:
            return json.JSONEncoder.default(self, obj)


class DataProcessor(object):
    fields = ('name', 'port', 'port_area', 'biz_type', 'size', 'charge_type', 'grade', 'price', 'effect_date',
              'price_type', 'oil_rate', 'is_sync_weight')

    # 无参构造函数
    def __init__(self):
        pass

    # 转成json
    def to_json(self):
        return json.dumps(self, default=lambda obj: obj.__dict__, indent=4, cls=JsonCustomEncoder)


# read xlsx file
# get workbook
sourceWorkbook = openpyxl.load_workbook('../files/鸭嘴兽应付费用导入模板.xlsx')
# get sheet by index
worksheet = sourceWorkbook.worksheets[0]
# get all rows
rows = worksheet.rows
# 遍历所有行
source_data_list = []
heads = []
for row_index, row in enumerate(rows):
    if row_index == 0:
        continue
    if row_index == 1:
        # get values from row
        heads = [cell.value for cell in row]
        continue
    # 遍历所有列, 带索引，从第3行开始
    data = DataProcessor()
    for index, cell in enumerate(row):
        if index == 1:
            data.name = cell.value
        elif index == 2:
            data.port = cell.value
        elif index == 3:
            data.port_area = cell.value
        elif index == 4:
            data.biz_type = cell.value
        elif index == 5:
            data.size = cell.value
        elif index == 6:
            data.charge_type = cell.value
        elif index == 7:
            data.grade = cell.value
        elif index == 8:
            if cell.value:
                data.price = str(cell.value)
        elif index == 9:
            # turn datetime to string
            if cell.value:
                data.effect_date = cell.value.strftime('%Y-%m-%d')
                continue
            data.effect_date = cell.value
        elif index == 10:
            data.price_type = cell.value
        elif index == 11:
            data.oil_rate = cell.value
        elif index == 12:
            data.is_sync_weight = cell.value
    source_data_list.append(data)

door_names_sheet = sourceWorkbook.worksheets[1]
door_names = []
# get all rows
door_names_rows = door_names_sheet.rows
# get value from first column every row
for row in door_names_rows:
    door_names.append(row[0].value)

# write source_data_list to new xlsx file
# Create a new workbook
new = openpyxl.Workbook()
# create new sheet
new_sheet = new.create_sheet('Sheet1', 0)
# write data to new sheet
# iterate door_names
new_sheet.append([])
new_sheet.append(heads)
# 前1000个数据
for door_name in door_names[:1000]:
    for index, data in enumerate(source_data_list):
        new_sheet.append(
            ["", door_name, data.port, data.port_area, data.biz_type, data.size, data.charge_type, data.grade,
             data.price, data.effect_date, data.price_type, data.oil_rate, data.is_sync_weight])
# 设置列宽
new_sheet.column_dimensions['A'].width = 20
new_sheet.column_dimensions['B'].width = 20
new_sheet.column_dimensions['C'].width = 20
new_sheet.column_dimensions['D'].width = 20
new_sheet.column_dimensions['E'].width = 20
new_sheet.column_dimensions['F'].width = 20
new_sheet.column_dimensions['G'].width = 20
new_sheet.column_dimensions['H'].width = 20
new_sheet.column_dimensions['I'].width = 20
new_sheet.column_dimensions['J'].width = 20
new_sheet.column_dimensions['K'].width = 20
new_sheet.column_dimensions['L'].width = 20
new_sheet.column_dimensions['M'].width = 20
new_sheet.column_dimensions['N'].width = 20
# save new workbook

new.save('../files/new.xlsx')

sourceWorkbook.close()
